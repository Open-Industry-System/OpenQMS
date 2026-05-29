# backend/app/utils/excel.py
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any
import urllib.parse
import zipfile

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse

# ─── 常量 ───
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

MAX_EXPORT_ROWS = 10000
MAX_IMPORT_ROWS = 5000
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB


def create_workbook(sheet_name: str, headers: list[str]) -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    return wb, ws


def append_row(ws: Worksheet, values: list[Any]) -> None:
    ws.append(values)


def auto_width(ws: Worksheet, min_width: int = 10, max_width: int = 40) -> None:
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(min(max_len + 2, max_width), min_width)


def workbook_to_bytes(wb: Workbook) -> bytes:
    auto_width(wb.active)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def excel_response(excel_bytes: bytes, filename: str) -> StreamingResponse:
    encoded = urllib.parse.quote(filename)
    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


def create_template(headers: list[str], sheet_name: str, example_row: list[Any] | None = None) -> bytes:
    wb, ws = create_workbook(sheet_name, headers)
    if example_row:
        ws.append(example_row)
    return workbook_to_bytes(wb)


@dataclass
class ImportError:
    row: int
    field: str
    message: str

@dataclass
class ImportResult:
    imported_count: int
    errors: list["ImportError"]


class ExcelParseError(Exception):
    pass


def parse_upload(
    file_bytes: bytes,
    header_mapping: dict[str, str],
    required_headers: list[str] | None = None,
    sheet_index: int = 0,
) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        wb = load_workbook(BytesIO(file_bytes), read_only=True)
    except (zipfile.BadZipFile, OSError, InvalidFileException) as e:
        raise ExcelParseError(f"文件格式无效，仅支持 .xlsx 格式: {e}")

    try:
        ws = wb.worksheets[sheet_index]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        col_map: dict[int, str] = {}
        matched_headers: set[str] = set()
        for col_idx, header in enumerate(header_row):
            if header:
                header_clean = str(header).strip().lower()
                for cn_header, internal_key in header_mapping.items():
                    if cn_header.strip().lower() == header_clean:
                        col_map[col_idx] = internal_key
                        matched_headers.add(cn_header.strip().lower())
                        break

        if required_headers:
            for req in required_headers:
                if req.strip().lower() not in matched_headers:
                    raise ExcelParseError(f"缺少必需表头：{req}")

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            record = {}
            all_none = True
            for col_idx, internal_key in col_map.items():
                value = row[col_idx] if col_idx < len(row) else None
                if value is not None:
                    # trim 字符串值，避免 "   " 通过必填校验
                    if isinstance(value, str):
                        value = value.strip()
                    if value is not None and value != "":
                        record[internal_key] = value
                        all_none = False
            if not all_none:
                rows.append({"_row": row_idx, **record})

        return rows
    finally:
        wb.close()


def coerce_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, float):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value)
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(s)
        except ValueError:
            return None
    return None


def coerce_int_strict(value: Any) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"必须为整数，不能为小数: {value}")
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s or '.' in s or not s.lstrip('-').isdigit():
            raise ValueError(f"必须为整数: {value}")
        return int(s)
    raise ValueError(f"无法转换为整数: {value}")
